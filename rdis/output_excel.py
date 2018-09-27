import datetime
import json
import openpyxl
import os
from collections import namedtuple
from log import log
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side

SAMPLE_PATH = '..\\..\\input\\simple_sample.txt'
# SAMPLE_PATH = '..\\..\\input\\easy.txt'
JSON_PATH = '..\\..\\output\\json\\公務資料.json'
# JSON_PATH = '..\\..\\output\\json\\json.json'
FOLDER_PATH = '..\\..\\output\\'+datetime.datetime.now().strftime('%Y%m%d_%H%M%S')+''

SAMPLE_TITLES = ['農戶編號', '調查姓名', '電話', '地址', '出生年', '原層別', '連結編號']
HOUSEHOLD_TITLES = ['[戶籍檔]', '姓名', '出生年', '關係', '農保', '老農津貼', '國民年金', '勞保給付', '勞退給付', '農保給付']
TRANSFER_CROP_TITLES = ['[轉作補貼]', '項目', '作物名稱', '金額', '期別']
DISASTER_TITLES = ['[災害]', '項目', '災害', '核定作物', '核定面積', '金額']
SB_SBDY_TITLES = ['[105小大]', '姓名', '災害', '大專業農轉契作', '小地主出租給付', '離農獎勵']
LIVESTOCK_TITLES = ['[畜牧資訊]', '年', '調查時間', '畜牧品項', '在養數量', '屠宰數量', '副產品名稱', '副產品數量']
SAMPLE_ROSTER_TITLES = ['序號', '樣本套號 ', '農戶編號', '連結編號 ', '戶長姓名', '電話 ', '地址 ', '層別 ', '經營種類 ', '可耕作地面積', '成功打勾']
SAMPLE_ATTR = [
        'layer',
        'name',
        'tel',
        'addr',
        'county',
        'town',
        'link_num',
        'id',
        'num',
        'main_type',
        'area',
        'sample_num',
    ]
Sample = namedtuple('Sample', SAMPLE_ATTR)

TYPE_FLAG = '主選'
ALIGNMENT = Alignment(horizontal='center', vertical='bottom')
SIDE =Side(style='medium')
BORDER = Border(
        top=SIDE,
        bottom=SIDE,
        left=SIDE,
        right=SIDE
    )

# sorted by county
sample_dict = {}
official_data = json.loads(open(JSON_PATH, encoding='utf8').read())

if not os.path.isdir(FOLDER_PATH):
    os.mkdir(FOLDER_PATH)

def set_excel_title(sheet, row_index, flag, *titles) -> None:
    if flag == 'sample_roster':
        for index, title in enumerate(titles[0], start=1):
            cell = sheet.cell(row_index, index)
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            cell.value = title
            cell.border = BORDER
    else:
        for index, title in enumerate(titles[0], start=1):
                sheet.cell(column=index, row=row_index).value = title


def read_sample() -> None:
    """
    讀取 sample 檔並使用 dict, key = county : value = 住在這縣市的人
    """
    with open(SAMPLE_PATH, encoding='utf8') as f:
        for line in f:
            sample = Sample._make(line.split('\t')) 
            county = sample.county
            if county not in sample_dict:
                county_l = []
                county_l.append(sample)
                sample_dict[county] = county_l
            else:
                sample_dict.get(county).append(sample)


def output_excel(type_flag=TYPE_FLAG) -> None:
    for county, samples in sample_dict.items():
        log.info('county : ' + county)
        if type_flag == '主選':
            samples.sort(key=lambda x:x.town)
        else:
            samples.sort(key=lambda x:x.num[-5:])
        wb = openpyxl.Workbook()
        col_index = 1
        row_index = 1
        county = county
        town = samples[0].town
        log.info('town : ' + town)
        sheet = wb.active
        sheet.title = town if type_flag == '主選' else 'sheet'+str(row_index+1)
        for sample in samples:
            log.info('person name : ' + sample.name)
            scholarship = ''
            sb = ''
            farmer_num = sample.num
            crops = []
            sample_data = official_data.get(farmer_num)
            if type_flag == '主選' and town != sample.town:
                town = sample.town
                sheet = wb.create_sheet(town)
                row_index = 1
            if row_index-1 == 0:
                width = list(map(lambda x: x*1.054,[14.29, 9.29, 16.29, 37.29, 9.29, 11.29, 11.29, 11.29, 11.29]))
                for i in range(1, len(width)+1):
                    sheet.column_dimensions[get_column_letter(i)].width = width[i-1]
            set_excel_title(sheet, row_index, 'sample', SAMPLE_TITLES)
            row_index += 1
            info = [
                farmer_num, sample_data.get('name'), sample_data.get('telephone'), sample_data.get('address'),
                sample_data.get('birthday'), sample_data.get('layer'), sample_data.get('serial')
            ]
            for index, value in enumerate(info, start=1):
                sheet.cell(column=index, row=row_index).value = value
                sheet.cell(column=index, row=row_index).alignment = Alignment(wrap_text=True)
            row_index += 1
            sheet.cell(column=col_index, row=row_index).value = ' ---------------------------------------------------------------- '
            row_index += 1
            set_excel_title(sheet, row_index, 'household', HOUSEHOLD_TITLES)
            household = sample_data.get('household')
            household.sort(key=lambda x: x[1])

            for person in household:
                row_index += 1
                for index, p_data in enumerate(person, start=2):
                    if index == 9:
                        scholarship += person[9]
                        continue
                    if index == 10 and person[10] not in sb:
                        sb += person[10]
                        break
                    sheet.cell(column=index, row=row_index).value = p_data
            # 輸出申報核定資料，檢查是否有資料
            declaration = sample_data.get('declaration')
            if declaration:
                row_index += 1
                sheet.cell(column=1, row=row_index).value = '[申報核定]'
                sheet.cell(column=2, row=row_index).value = declaration
            # 輸出轉作補貼資料，檢查是否有資料
            crop_sbdy = sample_data.get('cropSbdy')
            if crop_sbdy:
                crop_d = {}
                for i in crop_sbdy:
                    crop_name = i[0]
                    amount = int(i[1])
                    if crop_name not in crop_d:
                        crop_d[crop_name] = amount
                    else:
                        crop_d[crop_name] = crop_d.get(crop_name) + amount
                row_index += 1
                item_index = 0
                set_excel_title(sheet, row_index, 'transfer_crop', TRANSFER_CROP_TITLES)
                for k, v in crop_d.items():
                    row_index += 1
                    item_index += 1
                    sheet.cell(column=2, row=row_index).value = item_index
                    sheet.cell(column=3, row=row_index).value = k
                    sheet.cell(column=4, row=row_index).value = format(v, '8,d')
                    sheet.cell(column=5, row=row_index).value = '1'
                    if k not in crops:
                        crops.append(k)
            # 輸出災害補助資料，檢查是否有資料
            disaster = sample_data.get('disaster')
            if disaster:
                item_index = 0
                disaster_d = {}
                for i in disaster:
                    data = {}
                    disaster_name = i[0] + '-' + i[1]
                    area = float(i[2])
                    amount = int(i[3])
                    if disaster_name not in disaster_d:
                        data['area'] = area
                        data['amount'] = amount
                    else:
                        data = disaster_d.get(disaster_name)
                        data['area'] = data.get('area') + area
                        data['amount'] = data.get('amount') + amount
                    disaster_d[disaster_name] = data
                row_index += 1
                set_excel_title(sheet, row_index, 'disaster', DISASTER_TITLES)
                for k, v in disaster_d.items():
                    row_index += 1
                    item_index += 1
                    sheet.cell(column=2, row=row_index).value = item_index
                    l = k.split('-')
                    sheet.cell(column=3, row=row_index).value = l[0]
                    sheet.cell(column=4, row=row_index).value = l[1]
                    sheet.cell(column=5, row=row_index).value = v.get('area')
                    sheet.cell(column=6, row=row_index).value = format(v.get('amount'), '8,d')
                    if l[1] not in crops:
                        crops.append(l[1])
            # 輸出小大補助資料，檢查是否有資料
            sb_sbdy = sample_data.get('sbSbdy')
            if sb_sbdy:
                row_index += 1
                set_excel_title(sheet, row_index, 'sb_sbdy', SB_SBDY_TITLES)
                for i in sb_sbdy:
                    row_index += 1
                    for index, j in enumerate(i, start=1):
                        sheet.cell(column=index, row=row_index).value = j
                        
            # 輸出畜牧資料，檢查是否有資料
            livestock = sample_data.get('livestock')
            if livestock:
                row_index += 1
                set_excel_title(sheet, row_index, 'livestock', LIVESTOCK_TITLES)
                row_index += 1
                for k, v in livestock.items():
                    sheet.cell(column=1, row=row_index).value = k
                    v.sort(key=lambda x:x[6]+x[0])
                    for i in v:
                        sheet.cell(column=2, row=row_index).value = i[6]
                        sheet.cell(column=3, row=row_index).value = i[0]
                        sheet.cell(column=4, row=row_index).value = i[1]
                        sheet.cell(column=5, row=row_index).value = i[2]
                        sheet.cell(column=6, row=row_index).value = i[3]
                        sheet.cell(column=7, row=row_index).value = i[4]
                        sheet.cell(column=8, row=row_index).value = i[5]
                        row_index += 1
            else:
                row_index += 1
            # 輸出每月僱工資料
            mon_emp = sample_data.get('monEmp')
            titles = [
                        '[每月僱工]', '一月', '二月', '三月', '四月', '五月', '六月',
                        '七月', '八月', '九月', '十月', '十一月', '十二月'
                    ]
            for index, title in enumerate(titles, start=1):
                if index >= 8:
                    sheet.cell(column=index-6, row=row_index).value = title
                else:
                    sheet.cell(column=index, row=row_index).value = title
                if index == 7:
                    row_index += 1
                    for i, mon in enumerate(mon_emp, start=2):
                        sheet.cell(column=i, row=row_index).value = mon
                        if i == 7:
                            row_index += 1
                            break
                if index == 13:
                    row_index += 1
                    for i, mon in enumerate(mon_emp[6:], start=2):
                        sheet.cell(column=i, row=row_index).value = mon
            # 年度作物
            if len(crops) != 0:
                row_index += 1
                sheet.cell(column=1, row=row_index).value = '[105y-106y作物]'
                sheet.cell(column=2, row=row_index).value = ','.join(crops)
            # 小大與獎助學金
            if sb != '':
                row_index += 1
                sheet.cell(column=1, row=row_index).value = '[小大]'
                sheet.cell(column=2, row=row_index).value = sb
            
            if scholarship != '':
                row_index += 1
                sheet.cell(column=1, row=row_index).value = '[子女獎助學金]'
                sheet.cell(column=2, row=row_index).value = scholarship
            row_index += 1
            sheet.cell(column=1, row=row_index).value = ' ================================================================ '
        row_index += 1
        wb.save(FOLDER_PATH + '\\' + county + '.xlsx')
        output_sample_roster(county, samples)

# 輸出樣本名冊 excel
def output_sample_roster(c, s, type_flag=TYPE_FLAG) -> None:
    county = c
    town = s[0].town
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = town
    row_index = 4
    col_index = 1
    for sample in s:
        if town != sample.town:
            town = sample.town
            sheet = wb.create_sheet(town)
            row_index = 4
            col_index = 1
        if row_index == 4:
            width = list(map(lambda x: x*1.13,[5.29, 5.29, 13.29, 9.29, 9.29, 10.29, 50.29, 4.29, 5.29, 20.29, 5.29]))
            for i in range(1, len(width)+1):
                sheet.column_dimensions[get_column_letter(i)].width = width[i-1]
            titles = ['106年主力農家所得調查樣本名冊─'+type_flag, '本頁已完成調查戶數：_____', '失敗戶請填寫失敗訪視紀錄表', '']
            for index, title in enumerate(titles, start=1):
                sheet.merge_cells(start_row=index, start_column=col_index, end_row=index, end_column=11)
                cell = sheet.cell(index, col_index)
                cell.value = title
                cell.alignment = ALIGNMENT
                if index == 3:
                    cell.alignment = Alignment(horizontal='right')
                if index == 4:
                    for i in range(1, 12):
                        sheet.cell(index, i).border = BORDER
        sorted_sample = ['', sample.sample_num, sample.num, sample.num[-5:],
                         sample.name, sample.tel, sample.addr, sample.layer, sample.main_type, sample.area, '']
        row_index += 1
        set_excel_title(sheet, row_index, 'sample_roster', SAMPLE_ROSTER_TITLES)
        row_index += 1
        for index, i in enumerate(sorted_sample, start=1):
            cell = sheet.cell(row_index, index)
            if index in [2, 4, 8]:
                cell.alignment = ALIGNMENT 
            if index == 1:
                cell.value = row_index-5
            else:
                cell.value = i
            cell.border = BORDER
    wb.save(FOLDER_PATH + '\\' + county + '_樣本名冊.xlsx')
read_sample()
output_excel()